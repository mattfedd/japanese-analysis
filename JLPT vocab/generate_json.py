
from openpyxl import *
import json
import os

wb = load_workbook('vocab.xlsx')
ws = wb[wb.sheetnames[0]]

json_data = {}
output_file = "output.txt"
fp = open(output_file, "w")
fp.write('[')
for row in ws.iter_rows(min_row=1, min_col=1):
    json_data['JLPT'] = row[0].value
    json_data['kanji'] = row[1].value
    json_data['pronunciation_hira'] = row[2].value
    json_data['definition'] = row[3].value
    fp.write('\n')
    json.dump(json_data, fp)
    fp.write(',')
fp.seek(-1, os.SEEK_END)
fp.truncate()
fp.write("]")
